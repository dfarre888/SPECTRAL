#!/usr/bin/env python3
"""Idempotent A3DM RPAS xlsx → JSON export for Spectral + A3DM shared catalog.

Re-run when the spreadsheet updates:
  python3 scripts/import-a3dm-rpas.py [path/to/A3DM_RPAS_Database.xlsx]
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess

    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = ROOT / "data" / "a3dm" / "A3DM_RPAS_Database.xlsx"
OUT_DIR = ROOT / "data" / "a3dm"

SKIP_SPECTRUM_TYPES = {
    "Parachute Recovery",
    "Flight Termination",
    "Speaker",
    "Spotlight",
    "Strobe",
    "Dropper",
    "Gripper",
    "Tethering",
    "Beacon",
}

# Existing Spectral seed IDs — do not invent a second slug for these variants.
EXISTING_SLUGS = {
    "DRN-0029": "dji-mavic-3",
    "DRN-0084": "autel-evo-max-4t",
    "DRN-0099": "skydio-x10d",
}

MFR_SLUG = {
    "DJI": "dji",
    "Autel Robotics": "autel",
    "Skydio": "skydio",
    "Parrot": "parrot",
    "Yuneec": "yuneec",
    "Wingtra": "wingtra",
    "senseFly (AgEagle)": "sensefly",
    "Quantum Systems": "quantum",
    "Freefly Systems": "freefly",
    "JOUAV": "jouav",
    "XAG": "xag",
}


def slugify(text: str) -> str:
    s = text.lower().replace("&", " and ")
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-")


def drone_slug(drone_id: str, manufacturer: str, model: str, sub: str | None) -> str:
    if drone_id in EXISTING_SLUGS:
        return EXISTING_SLUGS[drone_id]
    mfr = MFR_SLUG.get(manufacturer, slugify(manufacturer.split("(")[0].strip()))
    base = slugify(model)
    if not base.startswith(mfr):
        base = f"{mfr}-{base}"
    if sub and sub not in ("Standard", "Original", None):
        sub_slug = slugify(str(sub))
        if sub_slug and sub_slug not in base:
            base = f"{base}-{sub_slug}"
    return base


def spectral_category(_a3dm_cat: str) -> str:
    """Shared catalog: all A3DM rows are COTS — airframe lives in a3dm_category."""
    return "cots"


def uas_group(a3dm_cat: str, mtow_g: float | None, dry_g: float | None) -> int:
    kg = None
    if mtow_g:
        kg = mtow_g / 1000.0
    elif dry_g:
        kg = dry_g / 1000.0
    c = (a3dm_cat or "").lower()
    if c == "fixed-wing":
        if kg and kg > 25:
            return 3
        return 2
    if c in ("heavy-lift", "agriculture") and kg and kg > 25:
        return 2
    if kg and kg > 25:
        return 2
    if kg and kg > 9:
        return 2
    return 1


def rows(ws):
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        yield dict(zip(headers, row))


def main() -> None:
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx.exists():
        raise SystemExit(f"xlsx not found: {xlsx}")

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)

    manufacturers = []
    for r in rows(wb["Manufacturers"]):
        manufacturers.append(
            {
                "id": r["manufacturer_id"],
                "name": r["name"],
                "country": r["country"],
                "type": r["type"],
                "website": r["website"],
                "notes": r["notes"],
            }
        )

    drones = []
    used_slugs: dict[str, str] = {}
    for r in rows(wb["Drones"]):
        slug = drone_slug(
            r["drone_id"],
            r["manufacturer_name"] or "",
            r["model"] or "",
            r["sub_category"],
        )
        if slug in used_slugs and used_slugs[slug] != r["drone_id"]:
            slug = f"{slug}-{r['drone_id'].lower()}"
        used_slugs[slug] = r["drone_id"]
        dry = r["dry_weight_g"]
        mtow = r["mtow_g"]
        payload = r["max_payload_g"]
        drones.append(
            {
                "a3dm_drone_id": r["drone_id"],
                "id": slug,
                "manufacturer_id": r["manufacturer_id"],
                "manufacturer": r["manufacturer_name"],
                "name": r["model"],
                "sub_category": r["sub_category"],
                "a3dm_category": r["category"],
                "category": spectral_category(r["category"] or ""),
                "dry_weight_g": dry,
                "mtow_g": mtow,
                "max_payload_g": payload,
                "year_released": r["year_released"],
                "notes": r["notes"],
                "uas_group": uas_group(r["category"] or "", mtow, dry),
            }
        )

    payloads = []
    for r in rows(wb["Payloads"]):
        ptype = r["type"] or "other"
        payloads.append(
            {
                "id": r["payload_id"],
                "manufacturer_id": r["manufacturer_id"],
                "manufacturer": r["manufacturer_name"],
                "name": r["name"],
                "type": ptype,
                "weight_g": r["weight_g"],
                "mount_type": r["mount_type"],
                "notes": r["notes"],
                "spectrum_eligible": ptype not in SKIP_SPECTRUM_TYPES,
            }
        )

    drone_by_a3dm = {d["a3dm_drone_id"]: d for d in drones}
    compatibility = []
    for r in rows(wb["Compatibility"]):
        drone = drone_by_a3dm.get(r["drone_id"])
        compatibility.append(
            {
                "id": r["compat_id"],
                "a3dm_drone_id": r["drone_id"],
                "platform_id": drone["id"] if drone else None,
                "payload_id": r["payload_id"],
                "drone_model": r["drone_model"],
                "payload_name": r["payload_name"],
                "notes": r["notes"],
            }
        )

    wb.close()

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    meta = {
        "source": "A3DM_RPAS_Database.xlsx",
        "manufacturers": len(manufacturers),
        "drones": len(drones),
        "payloads": len(payloads),
        "compatibility": len(compatibility),
        "skip_spectrum_types": sorted(SKIP_SPECTRUM_TYPES),
    }
    (OUT_DIR / "manufacturers.json").write_text(json.dumps(manufacturers, indent=2) + "\n")
    (OUT_DIR / "drones.json").write_text(json.dumps(drones, indent=2) + "\n")
    (OUT_DIR / "payloads.json").write_text(json.dumps(payloads, indent=2) + "\n")
    (OUT_DIR / "compatibility.json").write_text(json.dumps(compatibility, indent=2) + "\n")
    (OUT_DIR / "import-meta.json").write_text(json.dumps(meta, indent=2) + "\n")
    print(json.dumps(meta, indent=2))


if __name__ == "__main__":
    main()
